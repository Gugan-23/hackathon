import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
import webbrowser
import pyttsx3

# Function to speak text
def speak(text):
    engine = pyttsx3.init()
    engine.say(text)
    engine.runAndWait()

# Function to evaluate rule for order pairing
def evaluate_rule(rule, order1, order2):
    if rule == 1:
        # Rule #1: From the same kitchen, to the same customer, ready at the same time
        if order1['kitchen'] == order2['kitchen'] and order1['customer'] == order2['customer'] and abs(order1['ready_time'] - order2['ready_time']) <= 10:
            return 0
    elif rule == 2:
        # Rule #2: From two different kitchens, to the same customer, ready at the same time
        if order1['kitchen'] != order2['kitchen'] and order1['customer'] == order2['customer'] and abs(order1['ready_time'] - order2['ready_time']) <= 10:
            return 0
    elif rule == 3:
        # Rule #3: From the same kitchen, to two different customers, ready at the same time
        if order1['kitchen'] == order2['kitchen'] and order1['customer'] != order2['customer'] and abs(order1['ready_time'] - order2['ready_time']) <= 10:
            return 0
    elif rule == 4:
        # Rule #4: To the same customer, 2nd kitchens pick up on the way, ready at the time the rider reaches the second kitchen
        if order1['customer'] == order2['customer'] and order2['pickup_time'] <= order1['ready_time']:
            return 0
    elif rule == 5:
        # Rule #5: 2nd customers drop on the way to the 1st customer (Vice Versa), ready at the same time or by the time rider reaches the kitchen
        if (order1['customer'] != order2['customer'] and order2['pickup_time'] <= order1['ready_time']) or (order1['customer'] == order2['customer'] and abs(order1['ready_time'] - order2['ready_time']) <= 10):
            return 0
    elif rule == 6:
        # Rule #6: From the same kitchen, 2nd customers drop on the way to the customer 1st (Vice Versa), ready at the same time
        if order1['kitchen'] == order2['kitchen'] and order1['customer'] != order2['customer'] and abs(order1['ready_time'] - order2['ready_time']) <= 10:
            return 0
    return 1

# Function to generate all possible pairs of orders
def generate_order_pairs(orders):
    pairs = []
    for i in range(len(orders)):
        for j in range(i + 1, len(orders)):
            pairs.append((orders[i], orders[j]))
    return pairs

# Function to accept or reject each order pair based on the rules
def accept_order(order_pairs):
    for pair in order_pairs:
        scores = [0, 0, 0, 0, 0, 0]
        for i in range(1, 7):
            scores[i-1] = evaluate_rule(i, pair[0], pair[1])
        if scores.count(0) >= scores.count(1):
            print("Pair:", pair, "- Pick up the order")
        else:
            print("Pair:", pair, "- Do not pick up the order")

# Function to open the registration window
def open_registration_window():
    registration_window = tk.Toplevel()
    registration_window.title("Registration")

    # Function to save user details
    def save_user_details():
        # Get user details from entry fields
        username = username_entry.get()
        phone_number = phone_number_entry.get()
        location = location_entry.get()

        # Validate phone number
        if len(phone_number) != 10 or not phone_number.isdigit():
            messagebox.showerror("Error", "Phone number must be a 10-digit number.")
            return

        # Save user details to an Excel file
        try:
            wb = openpyxl.load_workbook(registered_user_path)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Registered Users"
            headers = ["Username", "Phone Number", "Location"]
            sheet.append(headers)

        data = [username, phone_number, location]
        sheet.append(data)

        wb.save(registered_user_path)

        # Display registration details
        messagebox.showinfo("Registration Details", f"Username: {username}\nPhone Number: {phone_number}\nLocation: {location}")

        # Enable access to customer and employee sections after registration
        customer_button.config(state=tk.NORMAL)
        employee_button.config(state=tk.NORMAL)
        
        # Speak registration success message
        speak("Registration successful.")

    # Create labels and entry fields for registration details
    username_label = tk.Label(registration_window, text="Username:")
    username_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
    username_entry = tk.Entry(registration_window)
    username_entry.grid(row=0, column=1, padx=5, pady=5)

    phone_number_label = tk.Label(registration_window, text="Phone Number:")
    phone_number_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
    phone_number_entry = tk.Entry(registration_window)
    phone_number_entry.grid(row=1, column=1, padx=5, pady=5)

    location_label = tk.Label(registration_window, text="Location:")
    location_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
    location_entry = tk.Entry(registration_window)
    location_entry.grid(row=2, column=1, padx=5, pady=5)

    # Create a button to save user details
    save_button = tk.Button(registration_window, text="Save", command=save_user_details)
    save_button.grid(row=3, columnspan=2, padx=5, pady=10)

# Function to open customer mode
def open_customer_mode():
    customer_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
    payment_status_button.grid(row=5, column=0, columnspan=2, padx=5, pady=10)
    employee_frame.grid_forget()

# Function to open employee mode
def open_employee_mode():
    employee_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
    payment_status_button.grid_forget()
    customer_frame.grid_forget()

# Function to provide payment status
def provide_payment_status(status):
    if status == "Success":
        messagebox.showinfo("Payment" "Status", "Order placed successfully.")
    else:
        messagebox.showwarning("Payment Status", "Payment failed.")

# Function to save customer order
def save_customer_order():
    customer_name = customer_name_entry.get()
    phone_number = phone_number_entry.get()
    food_choice = food_choice_var.get()
    user_location = location_entry.get()  # Get user's location

    if len(phone_number) != 10 or not phone_number.isdigit():
        messagebox.showerror("Error", "Phone number must be a 10-digit number.")
        return

    try:
        wb = openpyxl.load_workbook(order_file_path)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Order Details"
        headers = ["Customer Name", "Phone Number", "Food Choice", "Location"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=customer_name)
    sheet.cell(row=next_row, column=2, value=phone_number)
    sheet.cell(row=next_row, column=3, value=food_choice)
    sheet.cell(row=next_row, column=4, value=user_location)  # Save user's location with order

    wb.save(order_file_path)
    messagebox.showinfo("Success", "Order placed successfully.")

    webbrowser.open("https://gokuls-team-38.adalo.com/cookr-s-last-mile-delivery-1?_gl=1*1wo2019*_ga*MTEwNTA5MzUxMC4xNzEzMDg0Mzk0*_ga_SWT45DV35L*MTcxMzA5MTkxMS4yLjAuMTcxMzA5MTkxMS42MC4wLjA.")
    
    # Speak order placed message
    speak("Order placed successfully.")

# Function to save employee details
def save_employee_details():
    name = name_entry.get()
    age = age_entry.get()
    aadhar_number = aadhar_entry.get()
    address = address_entry.get()
    g_pay_number = g_pay_number_entry.get()
    phone_number = phone_number_entry.get()

    if not age.isdigit() or int(age) <= 18:
        messagebox.showerror("Error", "Age must be a valid number greater than 18.")
        return
    if len(aadhar_number) != 12 or not aadhar_number.isdigit():
        messagebox.showerror("Error", "Aadhar Card Number must be a 12-digit number.")
        return
    if len(g_pay_number) != 10 or not g_pay_number.isdigit():
        messagebox.showerror("Error", "G Pay Number must be a 10-digit number.")
        return
    if len(phone_number) != 10 or not phone_number.isdigit():
        messagebox.showerror("Error", "Phone Number must be a 10-digit number.")
        return

    wb = openpyxl.load_workbook(employee_file_path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value == aadhar_number:
                messagebox.showerror("Error", "User with this Aadhar Number already exists.")
                return

    next_row = sheet.max_row + 1
    details = [name, age, aadhar_number, address, g_pay_number, phone_number]
    for col, detail in enumerate(details, start=1):
        sheet.cell(row=next_row, column=col, value=detail)

    wb.save(employee_file_path)
    messagebox.showinfo("Success", "Employee details saved successfully.")
    
    # Speak employee details saved message
    speak("Employee details saved successfully.")

# Function to assign orders to employees
def assign_orders_to_employee():
    # Get location of order placed and assign employee of the same location
    try:
        wb = openpyxl.load_workbook(order_file_path)
        sheet = wb.active
    except FileNotFoundError:
        messagebox.showerror("Error", "Order details file not found.")
        return

    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            order_location = cell.value
            # Find the nearest employee based on order location and assign the order
            # You can implement this part based on your specific requirements

# Create the main GUI window
window = tk.Tk()
window.title("Order Management System")

# Load and resize the image
image_path = "C:\\Users\\chubr\\Downloads\\Cookr App logo.png"  # Replace with the path to your image
image = Image.open(image_path)
image = image.resize((window.winfo_screenwidth(), window.winfo_screenheight()))
background_image = ImageTk.PhotoImage(image)

# Set the background image
background_label = tk.Label(window, image=background_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

# File paths for Excel files
employee_file_path = "C:\\Users\\chubr\\OneDrive\\Documents\\cookr_employee_details.xlsx"
order_file_path = "C:\\Users\\chubr\\OneDrive\\Documents\Order_Details.xlsx"
registered_user_path = "C:\\Users\\chubr\\OneDrive\\Documents\\Registered_User.xlsx"

# Create a frame for the Employee section
employee_frame = tk.Frame(window, bd=2, relief=tk.GROOVE)

# Create labels and entry fields for employee details
name_label = tk.Label(employee_frame, text="Name:")
name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
name_entry = tk.Entry(employee_frame)
name_entry.grid(row=0, column=1, padx=5, pady=5)

age_label = tk.Label(employee_frame, text="Age:")
age_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
age_entry = tk.Entry(employee_frame)
age_entry.grid(row=1, column=1, padx=5, pady=5)

aadhar_label = tk.Label(employee_frame, text="Aadhar Card Number:")
aadhar_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
aadhar_entry = tk.Entry(employee_frame)
aadhar_entry.grid(row=2, column=1, padx=5, pady=5)

address_label = tk.Label(employee_frame, text="Address:")
address_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
address_entry = tk.Entry(employee_frame)
address_entry.grid(row=3, column=1, padx=5, pady=5)

g_pay_number_label = tk.Label(employee_frame, text="G Pay Number:")
g_pay_number_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
g_pay_number_entry = tk.Entry(employee_frame)
g_pay_number_entry.grid(row=4, column=1, padx=5, pady=5)

phone_number_label = tk.Label(employee_frame

, text="Phone Number:")
phone_number_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")
phone_number_entry = tk.Entry(employee_frame)
phone_number_entry.grid(row=5, column=1, padx=5, pady=5)

# Create a button to save employee details
save_employee_button = tk.Button(employee_frame, text="Save Employee", command=save_employee_details)
save_employee_button.grid(row=6, column=0, columnspan=2, padx=5, pady=10)

# Create a frame for the Customer section
customer_frame = tk.Frame(window, bd=2, relief=tk.GROOVE)

# Create labels and entry fields for customer details
customer_title_label = tk.Label(customer_frame, text="Order Details", font=("Arial", 14, "bold"))
customer_title_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

customer_name_label = tk.Label(customer_frame, text="Name:")
customer_name_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
customer_name_entry = tk.Entry(customer_frame)
customer_name_entry.grid(row=1, column=1, padx=5, pady=5)

phone_number_label = tk.Label(customer_frame, text="Phone Number:")
phone_number_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
phone_number_entry = tk.Entry(customer_frame)
phone_number_entry.grid(row=2, column=1, padx=5, pady=5)

food_choice_label = tk.Label(customer_frame, text="Food Choice:")
food_choice_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
food_choices = ["Veg", "Non-Veg"]
food_choice_var = tk.StringVar(value="Veg")
food_choice_dropdown = tk.OptionMenu(customer_frame, food_choice_var, *food_choices)
food_choice_dropdown.grid(row=3, column=1, padx=5, pady=5)

location_label = tk.Label(customer_frame, text="Location:")
location_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
location_entry = tk.Entry(customer_frame)
location_entry.grid(row=4, column=1, padx=5, pady=5)

# Create a button to save customer order
save_order_button = tk.Button(customer_frame, text="Place Order", command=save_customer_order)
save_order_button.grid(row=5, column=0, columnspan=2, padx=5, pady=10)

# Create buttons to switch between Customer and Employee modes
customer_button = tk.Button(window, text="Customer", command=open_customer_mode, state=tk.DISABLED)
customer_button.place(relx=0.85, rely=0.1, anchor=tk.CENTER)

employee_button = tk.Button(window, text="Employee", command=open_employee_mode, state=tk.DISABLED)
employee_button.place(relx=0.85, rely=0.2, anchor=tk.CENTER)

# Create a button for providing payment status
payment_status_button = tk.Button(customer_frame, text="Provide Payment Status", command=lambda: provide_payment_status("Success"))
payment_status_button.grid(row=6, column=0, columnspan=2, padx=5, pady=10)

# Create a button to open registration window
register_button = tk.Button(window, text="Register", command=open_registration_window)
register_button.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

# Start the GUI event loop
window.mainloop()
