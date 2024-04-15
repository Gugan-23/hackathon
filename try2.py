import tkinter as tk
from tkinter import messagebox
import openpyxl

class OrderManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Order Management System")

        # File paths for Excel files
        self.order_file_path = "C:\\Users\\chubr\\OneDrive\\Documents\\Order_Details.xlsx"
        self.employee_file_path = "C:\\Users\\chubr\\OneDrive\\Documents\\cookr_employee_details.xlsx"

        # Initialize orders list
        self.orders = []

        # Initialize GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Customer section
        customer_frame = tk.LabelFrame(self.root, text="Customer Section")
        customer_frame.grid(row=0, column=0, padx=10, pady=10)

        # Labels and entry fields for customer details
        customer_name_label = tk.Label(customer_frame, text="Customer Name:")
        customer_name_label.grid(row=0, column=0, padx=5, pady=5)
        self.customer_name_entry = tk.Entry(customer_frame)
        self.customer_name_entry.grid(row=0, column=1, padx=5, pady=5)

        kitchen_label = tk.Label(customer_frame, text="Kitchen:")
        kitchen_label.grid(row=1, column=0, padx=5, pady=5)
        self.kitchen_entry = tk.Entry(customer_frame)
        self.kitchen_entry.grid(row=1, column=1, padx=5, pady=5)

        ready_time_label = tk.Label(customer_frame, text="Ready Time:")
        ready_time_label.grid(row=2, column=0, padx=5, pady=5)
        self.ready_time_entry = tk.Entry(customer_frame)
        self.ready_time_entry.grid(row=2, column=1, padx=5, pady=5)

        pickup_time_label = tk.Label(customer_frame, text="Pickup Time:")
        pickup_time_label.grid(row=3, column=0, padx=5, pady=5)
        self.pickup_time_entry = tk.Entry(customer_frame)
        self.pickup_time_entry.grid(row=3, column=1, padx=5, pady=5)

        # Button to place order
        place_order_button = tk.Button(customer_frame, text="Place Order", command=self.place_order)
        place_order_button.grid(row=4, columnspan=2, padx=5, pady=10)

        # Employee section
        employee_frame = tk.LabelFrame(self.root, text="Employee Section")
        employee_frame.grid(row=0, column=1, padx=10, pady=10)

        # Button to pick up orders
        pick_up_orders_button = tk.Button(employee_frame, text="Pick Up Orders", command=self.pick_up_orders)
        pick_up_orders_button.grid(row=0, column=0, padx=5, pady=10)

        # Button for employee login
        employee_login_button = tk.Button(employee_frame, text="Employee Login", command=self.employee_login)
        employee_login_button.grid(row=1, column=0, padx=5, pady=10)

        # Button for employee registration
        employee_register_button = tk.Button(employee_frame, text="Register Employee", command=self.register_employee)
        employee_register_button.grid(row=2, column=0, padx=5, pady=10)

    def place_order(self):
        # Get customer details from entry fields
        customer_name = self.customer_name_entry.get()
        kitchen = self.kitchen_entry.get()
        ready_time = int(self.ready_time_entry.get())
        pickup_time = int(self.pickup_time_entry.get())

        # Validate ready time and pickup time
        if ready_time <= pickup_time:
            messagebox.showerror("Error", "Ready time must be greater than pickup time.")
            return

        # Add order to orders list
        self.orders.append({"customer": customer_name, "kitchen": kitchen, "ready_time": ready_time, "pickup_time": pickup_time})
        messagebox.showinfo("Success", "Order placed successfully.")

        # Save order details to Excel file
        self.save_order_details(customer_name, kitchen, ready_time, pickup_time)

    def save_order_details(self, customer_name, kitchen, ready_time, pickup_time):
        try:
            wb = openpyxl.load_workbook(self.order_file_path)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Order Details"
            headers = ["Customer Name", "Kitchen", "Ready Time", "Pickup Time"]
            sheet.append(headers)

        data = [customer_name, kitchen, ready_time, pickup_time]
        sheet.append(data)

        wb.save(self.order_file_path)

    def pick_up_orders(self):
        try:
            order_wb = openpyxl.load_workbook(self.order_file_path)
            order_sheet = order_wb.active
            orders = list(order_sheet.iter_rows(values_only=True))
            
            employee_wb = openpyxl.load_workbook(self.employee_file_path)
            employee_sheet = employee_wb.active
            employees = list(employee_sheet.iter_rows(values_only=True))

            # Initialize a dictionary to track the availability of employees
            available_employees = {employee[0]: True for employee in employees}

            # Iterate through orders
            for order in orders:
                order_kitchen = order[1]
                order_ready_time = order[2]
                order_assigned = False

                # Iterate through employees
                for employee in employees:
                    employee_name = employee[0]
                    employee_kitchen = employee[1]

                    # Check if employee belongs to the same kitchen and is available
                    if order_kitchen == employee_kitchen and available_employees[employee_name]:
                        # Assign order to employee
                        print(f"Order from {order_kitchen} kitchen assigned to employee {employee_name}")
                        order_assigned = True

                        # Mark employee as unavailable
                        available_employees[employee_name] = False

                        # Assuming each order is assigned to only one employee, break the loop
                        break
                
                # If order not assigned, display a message
                if not order_assigned:
                    print(f"No available employee found for order from {order_kitchen} kitchen.")

        except FileNotFoundError:
            messagebox.showerror("Error", "File not found.")

    def employee_login(self):
        # Prompt employee for login details
        login_window = tk.Toplevel(self.root)
        login_window.title("Employee Login")

        # Labels and entry fields for login details
        kitchen_label = tk.Label(login_window, text="Kitchen:")
        kitchen_label.grid(row=0, column=0, padx=5, pady=5)
        self.employee_kitchen_entry = tk.Entry(login_window)
        self.employee_kitchen_entry.grid(row=0, column=1, padx=5, pady=5)

        login_button = tk.Button(login_window, text="Login", command=self.assign_work)
        login_button.grid(row=1, column=0, columnspan=2, padx=5, pady=10)

    def register_employee(self):
        # Prompt employee for registration details
        registration_window = tk.Toplevel(self.root)
        registration_window.title("Employee Registration")

        # Labels and entry fields for registration details
        employee_name_label = tk.Label(registration_window, text="Employee Name:")
        employee_name_label.grid(row=0, column=0, padx=5, pady=5)
        self.employee_name_entry = tk.Entry(registration_window)
        self.employee_name_entry.grid(row=0, column=1, padx=5, pady=5)

        kitchen_label = tk.Label(registration_window, text="Kitchen:")
        kitchen_label.grid(row=1, column=0, padx=5, pady=5)
        self.employee_kitchen_entry_register = tk.Entry(registration_window)
        self.employee_kitchen_entry_register.grid(row=1, column=1, padx=5, pady=5)

        register_button = tk.Button(registration_window, text="Register", command=self.save_employee_details)
        register_button.grid(row=2, column=0, columnspan=2, padx=5, pady=10)

    def save_employee_details(self):
        # Get employee details from entry fields
        employee_name = self.employee_name_entry.get()
        kitchen = self.employee_kitchen_entry_register.get()

        try:
            wb = openpyxl.load_workbook(self.employee_file_path)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Employee Details"
            headers = ["Employee Name", "Kitchen"]
            sheet.append(headers)

        data = [employee_name, kitchen]
        sheet.append(data)

        wb.save(self.employee_file_path)

        messagebox.showinfo("Success", "Employee registered successfully.")

if __name__ == "__main__":
    root = tk.Tk()
    app = OrderManagementSystem(root)
    root.mainloop()
